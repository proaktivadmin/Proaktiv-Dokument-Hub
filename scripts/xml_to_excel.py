#!/usr/bin/env python3
"""
Convert Microsoft Excel XML Spreadsheet (.xml) files to .xlsx format.
"""

import argparse
import xml.etree.ElementTree as ET
from pathlib import Path

# Excel XML namespaces
SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"
NS = {
    "ss": SS_NS,
    "html": "http://www.w3.org/TR/REC-html40",
}


def _attr(element: ET.Element, name: str, default: str | None = None) -> str | None:
    """Get attribute by local name (handles namespace prefixes)."""
    for key, val in element.attrib.items():
        if key.endswith("}" + name) or key == name:
            return val
    return default


def get_text(element: ET.Element | None, default: str = "") -> str:
    """Extract text from a Data element inside a Cell."""
    if element is None:
        return default
    data = element.find("ss:Data", NS)
    if data is not None and data.text is not None:
        return data.text.strip()
    return default


def get_cell_value(cell: ET.Element) -> str | int | float:
    """Extract typed value from a Cell element."""
    data = cell.find("ss:Data", NS)
    if data is None:
        return ""
    dtype = _attr(data, "Type") or "String"
    if "String" in dtype:
        return (data.text or "").strip()
    if "Number" in dtype:
        raw = (data.text or "").strip()
        try:
            if "." in raw:
                return float(raw)
            return int(raw)
        except ValueError:
            return raw
    return (data.text or "").strip()


def parse_worksheet(worksheet: ET.Element) -> list[list[str | int | float]]:
    """Parse a Worksheet element into a list of rows (list of cell values)."""
    table = worksheet.find("ss:Table", NS)
    if table is None:
        return []

    rows_data: list[list[str | int | float]] = []
    for row in table.findall("ss:Row", NS):
        row_data: list[str | int | float] = []
        col_index = 0
        for cell in row.findall("ss:Cell", NS):
            # Handle Index attribute (sparse cells)
            idx = _attr(cell, "Index")
            if idx is not None:
                idx_int = int(idx)
                while len(row_data) < idx_int - 1:
                    row_data.append("")
                col_index = idx_int - 1

            value = get_cell_value(cell)
            merge_across = _attr(cell, "MergeAcross")
            if merge_across is not None:
                merge_count = int(merge_across)
                row_data.append(value)
                for _ in range(merge_count):
                    row_data.append("")
            else:
                row_data.append(value)
            col_index += 1
        rows_data.append(row_data)
    return rows_data


def convert_xml_to_xlsx(xml_path: Path, xlsx_path: Path) -> None:
    """Convert an Excel XML file to .xlsx format."""
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from None

    tree = ET.parse(xml_path)
    root = tree.getroot()
    workbook = root.find("ss:Workbook", NS)
    if workbook is None:
        workbook = root

    wb = openpyxl.Workbook()
    # Remove default sheet if we have multiple
    sheets = workbook.findall("ss:Worksheet", NS)
    default_sheet = wb.active

    for i, worksheet in enumerate(sheets):
        name = _attr(worksheet, "Name") or f"Sheet{i + 1}"
        # Excel sheet names: max 31 chars, no : \ / ? * [ ]
        safe_name = "".join(c for c in name[:31] if c not in ":\\/?*[]")

        if i == 0:
            ws = default_sheet
            ws.title = safe_name
        else:
            ws = wb.create_sheet(title=safe_name)

        rows_data = parse_worksheet(worksheet)
        for row_idx, row_data in enumerate(rows_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(xlsx_path)
    print(f"Converted: {xml_path.name} -> {xlsx_path.name}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Excel XML to .xlsx")
    parser.add_argument("files", nargs="+", type=Path, help="XML files to convert")
    parser.add_argument("-o", "--output-dir", type=Path, default=None, help="Output directory")
    args = parser.parse_args()

    for xml_path in args.files:
        if not xml_path.exists():
            print(f"Skip (not found): {xml_path}")
            continue
        if xml_path.suffix.lower() != ".xml" and ".xml" not in xml_path.suffix.lower():
            print(f"Skip (not XML): {xml_path}")
            continue

        out_dir = args.output_dir or xml_path.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        # Handle double .xml.xml extension
        stem = xml_path.stem
        if stem.endswith(".xml"):
            stem = stem[:-4]
        xlsx_path = out_dir / f"{stem}.xlsx"
        convert_xml_to_xlsx(xml_path, xlsx_path)


if __name__ == "__main__":
    main()
