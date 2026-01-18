"""
Proaktiv Photo Export Automation - UPDATED VERSION
Automates bulk photo exports from proaktiv.no/export to local storage

SELECTORS VERIFIED: January 18, 2026
- Form page: https://proaktiv.no/export
- Input field: First text input on page
- Submit button: First submit button on page

Usage:
    python photo_export_bot.py --excel "path/to/file.xlsx" --output "C:/Users/Adrian/Downloads"
    
Requirements:
    pip install playwright openpyxl
    playwright install chromium
"""

import asyncio
import argparse
import sys
from pathlib import Path
from datetime import datetime
import time
import shutil
import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout


class PhotoExportBot:
    def __init__(self, excel_file: str, output_dir: str, webdav_path: str = "proaktiv.no/shared"):
        self.excel_file = Path(excel_file)
        self.output_dir = Path(output_dir)
        self.webdav_path = Path(f"//{webdav_path}")  # Windows network path
        
        self.references = []
        self.results = {
            'success': [],
            'failed': [],
            'skipped': []
        }
        
    def extract_yellow_references(self) -> list[str]:
        """Extract property references from yellow-highlighted rows in Excel"""
        print(f"\n📋 Reading Excel file: {self.excel_file}")
        
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active
        
        yellow_refs = []
        for row_idx in range(2, sheet.max_row + 1):  # Skip header
            cell_b = sheet.cell(row=row_idx, column=2)  # Column B = Oppdragnr.
            
            # Check if yellow-highlighted
            if cell_b.fill and cell_b.fill.start_color and hasattr(cell_b.fill.start_color, 'rgb'):
                fill_color = cell_b.fill.start_color.rgb
                if 'FFFF' in str(fill_color).upper():  # Yellow = FFFFFF00
                    if cell_b.value:
                        yellow_refs.append(str(cell_b.value).strip())
        
        print(f"✅ Found {len(yellow_refs)} yellow-highlighted references")
        return yellow_refs
    
    async def submit_export(self, page, reference: str) -> bool:
        """Submit a reference to proaktiv.no/export"""
        try:
            print(f"  → Submitting reference: {reference}")
            
            # UPDATED SELECTORS (verified Jan 18, 2026)
            # Strategy: Find first text input and first submit button on page
            input_selector = 'input[type="text"]'
            button_selector = 'button[type="submit"]'
            
            # Clear any existing value and enter reference
            await page.fill(input_selector, '')
            await page.fill(input_selector, reference)
            await page.wait_for_timeout(500)  # Brief pause
            
            # Submit the form
            await page.click(button_selector)
            
            # Wait for response (either success or error)
            await page.wait_for_timeout(3000)  # Wait 3 seconds for server response
            
            # Check for error messages (common error selectors)
            error_selectors = [
                '.error',
                '.alert-danger',
                '[role="alert"]',
                '.message.error',
                '.notification.error'
            ]
            
            for selector in error_selectors:
                error = await page.query_selector(selector)
                if error:
                    error_text = await error.text_content()
                    print(f"  ❌ Export failed: {error_text}")
                    return False
            
            print(f"  ✅ Export submitted successfully")
            return True
            
        except Exception as e:
            print(f"  ❌ Error submitting: {str(e)}")
            return False
    
    async def wait_for_webdav_folder(self, reference: str, timeout: int = 180) -> bool:
        """Poll WebDAV location for folder to appear"""
        folder_path = self.webdav_path / reference
        
        print(f"  ⏳ Waiting for folder in WebDAV (timeout: {timeout}s)")
        print(f"     Looking for: {folder_path}")
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            if folder_path.exists():
                # Verify folder has files
                try:
                    files = list(folder_path.glob('*'))
                    if files:
                        print(f"  ✅ Folder ready with {len(files)} files")
                        return True
                except Exception as e:
                    print(f"  ⚠️  Error accessing folder: {e}")
            
            await asyncio.sleep(5)  # Check every 5 seconds
            elapsed = int(time.time() - start_time)
            if elapsed % 30 == 0:  # Progress update every 30 seconds
                print(f"  ⏳ Still waiting... ({elapsed}s elapsed)")
        
        print(f"  ❌ Timeout: Folder did not appear after {timeout}s")
        return False
    
    def copy_to_local(self, reference: str) -> bool:
        """Copy folder from WebDAV to local output directory"""
        source = self.webdav_path / reference
        destination = self.output_dir / reference
        
        try:
            print(f"  📁 Copying from: {source}")
            print(f"  📁 Copying to: {destination}")
            
            if destination.exists():
                print(f"  ⚠️  Destination exists, removing old version")
                shutil.rmtree(destination)
            
            shutil.copytree(source, destination)
            
            # Verify copy
            files = list(destination.glob('*'))
            print(f"  ✅ Copied {len(files)} files successfully")
            return True
            
        except Exception as e:
            print(f"  ❌ Copy failed: {str(e)}")
            return False
    
    async def process_reference(self, page, reference: str) -> bool:
        """Process a single reference: submit, wait, copy"""
        print(f"\n{'='*60}")
        print(f"Processing: {reference}")
        print(f"{'='*60}")
        
        # Step 1: Submit export
        if not await self.submit_export(page, reference):
            self.results['failed'].append(reference)
            return False
        
        # Step 2: Wait for WebDAV folder
        if not await self.wait_for_webdav_folder(reference):
            self.results['failed'].append(reference)
            return False
        
        # Step 3: Copy to local
        if not self.copy_to_local(reference):
            self.results['failed'].append(reference)
            return False
        
        self.results['success'].append(reference)
        return True
    
    async def run(self):
        """Main execution flow"""
        print("\n" + "="*60)
        print("🚀 Proaktiv Photo Export Automation")
        print("="*60)
        
        # Step 1: Extract references
        self.references = self.extract_yellow_references()
        
        if not self.references:
            print("❌ No yellow-highlighted references found!")
            return
        
        print(f"\n📊 Summary:")
        print(f"  • References to export: {len(self.references)}")
        print(f"  • Estimated time: {len(self.references) * 1.5:.0f}-{len(self.references) * 2:.0f} minutes")
        print(f"  • Output directory: {self.output_dir}")
        print(f"  • WebDAV path: {self.webdav_path}")
        
        # Show all references
        print(f"\n📋 References to export:")
        for i, ref in enumerate(self.references, 1):
            print(f"  {i:2d}. {ref}")
        
        # Confirmation
        print("\n⚠️  BEFORE PROCEEDING:")
        print("  1. Make sure you are logged into https://proaktiv.no/export")
        print("  2. Verify WebDAV is accessible: Open File Explorer → \\\\proaktiv.no\\shared")
        print("  3. Close any Excel files that might be using the references")
        
        response = input("\n✅ Ready to start? (y/n): ")
        if response.lower() != 'y':
            print("❌ Cancelled by user")
            return
        
        # Create output directory
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Step 2: Browser automation
        async with async_playwright() as p:
            print("\n🌐 Launching browser...")
            browser = await p.chromium.launch(headless=False)  # headless=False to see what's happening
            context = await browser.new_context()
            page = await context.new_page()
            
            # Navigate to export page
            print("  → Navigating to https://proaktiv.no/export")
            await page.goto('https://proaktiv.no/export')
            await page.wait_for_load_state('networkidle')
            
            # Verify login
            print("\n⏸️  MANUAL CHECK:")
            print("  Please verify you are logged in and can see the export form.")
            input("  Press Enter when ready to continue...")
            
            # Process each reference
            start_time = time.time()
            for i, reference in enumerate(self.references, 1):
                print(f"\n{'#'*60}")
                print(f"# [{i}/{len(self.references)}] Processing: {reference}")
                print(f"{'#'*60}")
                
                try:
                    await self.process_reference(page, reference)
                except Exception as e:
                    print(f"❌ Unexpected error: {str(e)}")
                    self.results['failed'].append(reference)
                
                # Progress update
                elapsed = time.time() - start_time
                avg_time = elapsed / i
                remaining = (len(self.references) - i) * avg_time
                print(f"\n⏱️  Progress: {i}/{len(self.references)} | Elapsed: {elapsed/60:.1f}m | Est. remaining: {remaining/60:.1f}m")
                
                # Brief pause between exports
                if i < len(self.references):
                    print(f"  ⏸️  Waiting 5 seconds before next export...")
                    await asyncio.sleep(5)
            
            await browser.close()
        
        # Final summary
        self.print_summary()
    
    def print_summary(self):
        """Print final results summary"""
        print("\n" + "="*60)
        print("📊 EXPORT SUMMARY")
        print("="*60)
        
        print(f"\n✅ Successful: {len(self.results['success'])}/{len(self.references)}")
        for ref in self.results['success']:
            print(f"  • {ref}")
        
        if self.results['failed']:
            print(f"\n❌ Failed: {len(self.results['failed'])}")
            for ref in self.results['failed']:
                print(f"  • {ref}")
        
        if self.results['skipped']:
            print(f"\n⏭️  Skipped: {len(self.results['skipped'])}")
            for ref in self.results['skipped']:
                print(f"  • {ref}")
        
        # Save results to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = self.output_dir / f"export_report_{timestamp}.txt"
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("PROAKTIV PHOTO EXPORT REPORT\n")
            f.write("="*60 + "\n\n")
            f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total references: {len(self.references)}\n")
            f.write(f"Successful: {len(self.results['success'])}\n")
            f.write(f"Failed: {len(self.results['failed'])}\n\n")
            
            f.write("SUCCESSFUL EXPORTS:\n")
            for ref in self.results['success']:
                f.write(f"  • {ref}\n")
            
            if self.results['failed']:
                f.write("\nFAILED EXPORTS:\n")
                for ref in self.results['failed']:
                    f.write(f"  • {ref}\n")
        
        print(f"\n📄 Report saved to: {report_file}")
        print("\n✨ Export automation complete!")


def main():
    parser = argparse.ArgumentParser(description='Automate photo exports from proaktiv.no')
    parser.add_argument('--excel', required=True, help='Path to Excel file with references')
    parser.add_argument('--output', default='./downloads', help='Output directory for photos')
    parser.add_argument('--webdav', default='proaktiv.no/shared', help='WebDAV network path')
    
    args = parser.parse_args()
    
    bot = PhotoExportBot(
        excel_file=args.excel,
        output_dir=args.output,
        webdav_path=args.webdav
    )
    
    asyncio.run(bot.run())


if __name__ == '__main__':
    main()
