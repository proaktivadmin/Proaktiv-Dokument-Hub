"""
Sales Report Service

Exports broker revenue (vederlag + andre inntekter) for sold properties
via Vitec Hub API. Matches scope: Proaktiv Eiendomsmegling AS (1120),
statuses 40-48 (sold/completed), Hovedbokskonti for vederlag and andre inntekter.

Timezone policy
---------------
All report periods use **Europe/Oslo** for business semantics (month
boundaries, fiscal year cutoffs).  Timestamps stored in PostgreSQL are
**UTC with timezone** (``DateTime(timezone=True)``).  Conversion happens at
the boundary: user inputs are interpreted as Oslo dates, the service
converts to UTC for storage and queries, and display values convert back.

Multi-source architecture
-------------------------
Cache tables carry a ``data_source`` discriminator so historical imports
and future KPI sources coexist with live Vitec Next data.  Current values:
``vitec_next`` (live API sync) and ``legacy_import`` (deferred).
"""

from __future__ import annotations

import asyncio
import hashlib
import io
import logging
import re
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo

from sqlalchemy import and_, select, text
from sqlalchemy import func as sa_func

from app.config import settings
from app.database import async_session_factory
from app.models.office import Office
from app.models.report_sales_cache import (
    DATA_SOURCE_VITEC_NEXT,
    ReportSalesCacheState,
    ReportSalesEstateCache,
    ReportSalesSyncEvent,
    ReportSalesTransactionCache,
)
from app.services.vitec_hub_service import VitecHubService

logger = logging.getLogger(__name__)

try:
    REPORT_TIMEZONE = ZoneInfo("Europe/Oslo")
except KeyError:
    from datetime import timezone as _tz

    REPORT_TIMEZONE = _tz(timedelta(hours=1))  # CET fallback when tzdata is missing
    logger.debug("tzdata package not installed; using fixed UTC+1 fallback for REPORT_TIMEZONE")

# Hovedbokskonti for vederlag (remuneration) - from HOVEDBOKSKONTI UI
VEDERLAG_ACCOUNTS = {
    "3000",
    "3001",
    "3002",
    "3003",
    "3006",
    "3009",
    "3010",
    "3013",
    "3015",
    "3019",
    "3111",
    "3112",
    "3113",
    "3115",
    "3220",
    "3221",
}

# Hovedbokskonti for andre inntekter (other income)
ANDRE_INNTEKTER_ACCOUNTS = {
    "3005",
    "3018",
    "3020",
    "3030",
    "3031",
    "3050",
    "8050",
}

REVENUE_ACCOUNTS = VEDERLAG_ACCOUNTS | ANDRE_INNTEKTER_ACCOUNTS

# Default department: Proaktiv Eiendomsmegling AS
DEFAULT_DEPARTMENT_ID = 1120

SYNC_SOURCES: dict[str, dict[str, Any]] = {
    "vitec_next": {
        "label": "Vitec Next",
        "data_source": DATA_SOURCE_VITEC_NEXT,
        "accounts": REVENUE_ACCOUNTS,
        "account_categories": {
            "vederlag": sorted(VEDERLAG_ACCOUNTS),
            "andre_inntekter": sorted(ANDRE_INNTEKTER_ACCOUNTS),
        },
        "coverage_start": "2026-02",
    },
}


def _parse_iso_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt
    except (ValueError, TypeError):
        return None


def _month_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _iter_months(from_dt: datetime, to_dt: datetime) -> list[tuple[int, int]]:
    start = datetime(from_dt.year, from_dt.month, 1, tzinfo=from_dt.tzinfo or timezone.utc)
    end = datetime(to_dt.year, to_dt.month, 1, tzinfo=to_dt.tzinfo or timezone.utc)
    out: list[tuple[int, int]] = []
    cursor = start
    while cursor <= end:
        out.append((cursor.year, cursor.month))
        if cursor.month == 12:
            cursor = cursor.replace(year=cursor.year + 1, month=1)
        else:
            cursor = cursor.replace(month=cursor.month + 1)
    return out


def _month_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
    if month == 12:
        next_month = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        next_month = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    end = next_month - timedelta(seconds=1)
    return start, end


def _normalize_account(account: str | None) -> str:
    """Normalize account number for comparison (strip, handle None)."""
    if account is None:
        return ""
    return str(account).strip()


def _format_date_iso(iso_str: str | None) -> str:
    """Format ISO date to dd.mm.yyyy for display."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        return str(iso_str)


def _looks_like_uuid(s: str) -> bool:
    """True if string looks like a UUID (e.g. estate_id fallback)."""
    if not s or len(s) < 30:
        return False
    return bool(re.match(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$", s.strip()))


def _build_estate_address(est: dict) -> str:
    """Build display address from estate object (Vitec may use address, streetAddress, etc.)."""
    addr = str(est.get("address") or "").strip()
    if addr:
        return addr
    street = str(est.get("streetAddress") or est.get("street_address") or "").strip()
    postal = str(est.get("postalCode") or est.get("zipCode") or est.get("postal_code") or "").strip()
    city = str(est.get("city") or est.get("visitCity") or "").strip()
    parts = [p for p in [street, f"{postal} {city}".strip()] if p]
    return ", ".join(parts) if parts else ""


def _revenue_amount(amount: float, vat_amount: float, include_vat: bool) -> float:
    """Revenue as positive number (API may return negative for credits)."""
    total = float(amount) + (float(vat_amount) if include_vat else 0)
    return abs(total)


def _extract_display_value(obj: object) -> str:
    """Extract display string from API value (string, or dict with name/key)."""
    if obj is None:
        return ""
    if isinstance(obj, str):
        return obj.strip()
    if isinstance(obj, dict):
        return str(obj.get("name") or obj.get("key") or obj.get("value") or obj.get("displayName") or "").strip()
    return str(obj).strip()


def _build_estate_metadata(est: dict) -> dict[str, str]:
    """Extract property type, assignment type, and oppdragsnummer from estate (Vitec field names may vary)."""
    prop_type = _extract_display_value(
        est.get("propertyType")
        or est.get("property_type")
        or est.get("estateType")
        or est.get("estate_type")
        or est.get("boligtype")
        or est.get("grunntype")
    )
    assign_type = _extract_display_value(
        est.get("assignmentType") or est.get("assignment_type") or est.get("oppdragstype")
    )
    oppdrag = _extract_display_value(
        est.get("assignmentNumber")
        or est.get("assignment_number")
        or est.get("oppdragsnummer")
        or est.get("oppdragsnr")
    )
    return {
        "property_type": prop_type or "—",
        "assignment_type": assign_type or "—",
        "assignment_number": oppdrag or "",
    }


def _infer_broker_role(employee: dict) -> str:
    """Infer performer category from employee fields."""
    role_blob = " ".join(
        [
            str(employee.get("title") or ""),
            str(employee.get("jobTitle") or ""),
            str(employee.get("employeeType") or ""),
            str(employee.get("role") or ""),
        ]
    ).lower()
    if "fullmektig" in role_blob:
        return "eiendomsmeglerfullmektig"
    if "eiendomsmegler" in role_blob or re.search(r"\bmegler\b", role_blob):
        return "eiendomsmegler"
    return "unknown"


class SalesReportService:
    """Build sales report from Vitec Hub Accounting API."""

    def __init__(self, hub: VitecHubService | None = None) -> None:
        self._hub = hub or VitecHubService()

    async def build_report(
        self,
        *,
        department_id: int = DEFAULT_DEPARTMENT_ID,
        year: int | None = None,
        from_date: str | None = None,
        to_date: str | None = None,
        include_vat: bool = False,
    ) -> bytes:
        """
        Build Excel sales report for brokers who sold properties this year.
        """
        result = await self._fetch_report_data(
            department_id=department_id,
            year=year,
            from_date=from_date,
            to_date=to_date,
            include_vat=include_vat,
        )
        report_data, broker_map, brokers_with_sales, broker_estates, estate_address, estate_metadata, _ = result
        y = report_data["year"]
        from_date = report_data["from_date"]
        to_date = report_data["to_date"]

        return self._build_excel(
            year=y,
            department_id=department_id,
            from_date=from_date,
            to_date=to_date,
            broker_map=broker_map,
            brokers_with_sales=brokers_with_sales,
            broker_estates=broker_estates,
            estate_address=estate_address,
            estate_metadata=estate_metadata,
            include_vat=include_vat,
        )

    async def get_report_data(
        self,
        *,
        department_id: int = DEFAULT_DEPARTMENT_ID,
        year: int | None = None,
        from_date: str | None = None,
        to_date: str | None = None,
        include_vat: bool = False,
    ) -> dict:
        """
        Fetch and return sales report data as structured dict for JSON API.
        """
        result = await self._fetch_report_data(
            department_id=department_id,
            year=year,
            from_date=from_date,
            to_date=to_date,
            include_vat=include_vat,
        )
        return result[0]

    async def get_franchise_report_data(
        self,
        *,
        year: int | None = None,
        from_date: str | None = None,
        to_date: str | None = None,
        include_vat: bool = False,
        department_ids: list[int] | None = None,
    ) -> dict:
        """
        Fetch franchise report data across departments.
        """
        installation_id = settings.VITEC_INSTALLATION_ID
        if not installation_id:
            raise ValueError("VITEC_INSTALLATION_ID is not configured.")

        if department_ids:
            selected_departments = list(dict.fromkeys(department_ids))
        else:
            departments_raw = await self._hub.get_departments(installation_id)
            selected_departments = []
            for dep in departments_raw or []:
                dep_id = dep.get("departmentId") or dep.get("id")
                dep_name = str(dep.get("name") or "").lower()
                if dep_id is None:
                    continue
                if "oppgjør" in dep_name or "pacta" in dep_name:
                    continue
                try:
                    selected_departments.append(int(dep_id))
                except (TypeError, ValueError):
                    continue

        selected_departments = list(dict.fromkeys(selected_departments))
        if not selected_departments:
            return {
                "year": year or datetime.now().year,
                "from_date": from_date,
                "to_date": to_date,
                "include_vat": include_vat,
                "departments": [],
                "summary": {
                    "total_sales": 0,
                    "total_revenue": 0.0,
                    "department_count": 0,
                },
            }

        tasks = [
            self.get_report_data(
                department_id=dep_id,
                year=year,
                from_date=from_date,
                to_date=to_date,
                include_vat=include_vat,
            )
            for dep_id in selected_departments
        ]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        dept_name_map = await self._resolve_department_names(selected_departments)

        departments: list[dict] = []
        total_sales = 0
        total_revenue = 0.0

        for dep_id, result in zip(selected_departments, results, strict=False):
            if isinstance(result, Exception):
                logger.warning("Skipping franchise department %s: %s", dep_id, result)
                continue
            dep_revenue = float(result.get("total_revenue") or 0.0)
            dep_sales = int(result.get("total_sales") or 0)
            total_sales += dep_sales
            total_revenue += dep_revenue
            departments.append(
                {
                    "department_id": dep_id,
                    "department_name": dept_name_map.get(dep_id, f"Avdeling {dep_id}"),
                    "total_sales": dep_sales,
                    "total_revenue": round(dep_revenue, 2),
                    "brokers": result.get("brokers", []),
                }
            )

        for dep in departments:
            dep_total = float(dep["total_revenue"])
            dep["revenue_share_percent"] = round((dep_total / total_revenue * 100.0), 2) if total_revenue > 0 else 0.0

        departments.sort(key=lambda d: d["total_revenue"], reverse=True)
        first = departments[0] if departments else {}
        return {
            "year": first.get("year", year or datetime.now().year),
            "from_date": first.get("from_date", from_date),
            "to_date": first.get("to_date", to_date),
            "from_date_display": first.get("from_date_display"),
            "to_date_display": first.get("to_date_display"),
            "include_vat": include_vat,
            "departments": departments,
            "summary": {
                "total_sales": total_sales,
                "total_revenue": round(total_revenue, 2),
                "department_count": len(departments),
            },
        }

    async def get_best_performers_data(
        self,
        *,
        year: int | None = None,
        from_date: str | None = None,
        to_date: str | None = None,
        include_vat: bool = False,
        department_ids: list[int] | None = None,
        top_n: int = 5,
    ) -> dict:
        """
        Build best-performers leaderboard by role and department.
        """
        franchise = await self.get_franchise_report_data(
            year=year,
            from_date=from_date,
            to_date=to_date,
            include_vat=include_vat,
            department_ids=department_ids,
        )

        installation_id = settings.VITEC_INSTALLATION_ID
        employees = await self._hub.get_employees(installation_id) if installation_id else []
        role_by_broker: dict[str, str] = {}
        for emp in employees or []:
            eid = str(emp.get("employeeId") or "").strip()
            if not eid:
                continue
            role_by_broker[eid] = _infer_broker_role(emp)

        brokers_agg: dict[str, dict] = defaultdict(
            lambda: {
                "broker_id": "",
                "name": "",
                "role": "unknown",
                "total_revenue": 0.0,
                "total_sales": 0,
                "properties_by_estate": {},
            }
        )
        for dep in franchise.get("departments", []):
            for broker in dep.get("brokers", []):
                bid = str(broker.get("broker_id") or "").strip()
                if not bid:
                    continue
                row = brokers_agg[bid]
                row["broker_id"] = bid
                row["name"] = broker.get("name") or bid
                row["role"] = role_by_broker.get(bid, "unknown")
                row["total_revenue"] += float(broker.get("total") or 0.0)
                row["total_sales"] += int(broker.get("sale_count") or 0)
                for prop in broker.get("properties") or []:
                    eid = str(prop.get("estate_id") or "").strip()
                    if not eid:
                        continue
                    if eid not in row["properties_by_estate"]:
                        row["properties_by_estate"][eid] = prop

        def _sort_rows(rows: list[dict]) -> list[dict]:
            return sorted(rows, key=lambda r: (-float(r["total_revenue"]), str(r["name"]).lower()))

        def _row_with_properties(r: dict) -> dict:
            props = sorted(
                r["properties_by_estate"].values(),
                key=lambda p: (str(p.get("address") or ""), str(p.get("estate_id") or "")),
            )
            return {
                "broker_id": r["broker_id"],
                "name": r["name"],
                "role": r["role"],
                "total_revenue": round(float(r["total_revenue"]), 2),
                "total_sales": r["total_sales"],
                "properties": props,
            }

        rows = [_row_with_properties(r) for r in brokers_agg.values()]
        eiendomsmegler = _sort_rows([r for r in rows if r["role"] == "eiendomsmegler"])[:top_n]
        fullmektig = _sort_rows([r for r in rows if r["role"] == "eiendomsmeglerfullmektig"])[:top_n]
        unknown = _sort_rows([r for r in rows if r["role"] == "unknown"])[:top_n]
        dept_list = sorted(
            franchise.get("departments", []),
            key=lambda d: (-float(d.get("total_revenue") or 0.0), str(d.get("department_name", ""))),
        )[:top_n]
        departments = [
            {
                "department_id": d["department_id"],
                "department_name": d["department_name"],
                "total_revenue": d["total_revenue"],
                "total_sales": d["total_sales"],
                "brokers": d.get("brokers", []),
            }
            for d in dept_list
        ]

        return {
            "from_date": franchise.get("from_date"),
            "to_date": franchise.get("to_date"),
            "from_date_display": franchise.get("from_date_display"),
            "to_date_display": franchise.get("to_date_display"),
            "include_vat": include_vat,
            "top_n": top_n,
            "eiendomsmegler": eiendomsmegler,
            "eiendomsmeglerfullmektig": fullmektig,
            "unknown": unknown,
            "departments": departments,
        }

    async def list_cache_events(
        self,
        *,
        department_id: int | None = None,
        since_id: int | None = None,
        limit: int = 100,
    ) -> list[ReportSalesSyncEvent]:
        installation_id = settings.VITEC_INSTALLATION_ID or ""
        async with async_session_factory() as db:
            stmt = select(ReportSalesSyncEvent).where(ReportSalesSyncEvent.installation_id == installation_id)
            if department_id is not None:
                stmt = stmt.where(ReportSalesSyncEvent.department_id == department_id)
            if since_id is not None:
                stmt = stmt.where(ReportSalesSyncEvent.id > since_id)
            stmt = stmt.order_by(ReportSalesSyncEvent.id.desc()).limit(max(1, min(limit, 500)))
            rows = list((await db.execute(stmt)).scalars().all())
            return list(reversed(rows))

    async def build_best_performers_report(
        self,
        *,
        year: int | None = None,
        from_date: str | None = None,
        to_date: str | None = None,
        include_vat: bool = False,
        department_ids: list[int] | None = None,
        top_n: int = 5,
    ) -> bytes:
        """
        Generate Excel workbook for best performers leaderboard.
        """
        import openpyxl
        from openpyxl.styles import Font

        data = await self.get_best_performers_data(
            year=year,
            from_date=from_date,
            to_date=to_date,
            include_vat=include_vat,
            department_ids=department_ids,
            top_n=top_n,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Best performers"
        ws.append(["Best performers"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Periode: {data.get('from_date_display', '')} - {data.get('to_date_display', '')}"])
        ws.append([])

        def add_section(title: str, rows: list[dict], name_key: str = "name") -> None:
            ws.append([title])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.append(["Navn", "Antall salg", "Omsetning (kr)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
            for row in rows:
                ws.append(
                    [
                        row.get(name_key, "—"),
                        int(row.get("total_sales", 0)),
                        round(float(row.get("total_revenue", 0.0)), 0),
                    ]
                )
            ws.append([])

        add_section("Eiendomsmegler", data.get("eiendomsmegler", []))
        add_section("Eiendomsmeglerfullmektig", data.get("eiendomsmeglerfullmektig", []))
        if data.get("unknown"):
            add_section("Ukjent rolle", data.get("unknown", []))
        add_section("Avdelinger", data.get("departments", []), name_key="department_name")

        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def _fetch_report_data(
        self,
        *,
        department_id: int = DEFAULT_DEPARTMENT_ID,
        year: int | None = None,
        from_date: str | None = None,
        to_date: str | None = None,
        include_vat: bool = False,
    ) -> tuple:
        """Fetch report data and return (report_data_dict, broker_map, ...) for Excel or JSON."""
        if not self._hub.is_configured:
            raise ValueError("Vitec Hub credentials are not configured.")

        installation_id = settings.VITEC_INSTALLATION_ID
        if not installation_id:
            raise ValueError("VITEC_INSTALLATION_ID is not configured.")

        y = year or datetime.now().year
        if from_date:
            from_date = from_date if "T" in from_date else f"{from_date}T00:00:00"
        else:
            from_date = f"{y}-01-01T00:00:00"
        if to_date:
            to_date = to_date if "T" in to_date else f"{to_date}T23:59:59"
        else:
            to_date = datetime.now().strftime("%Y-%m-%dT23:59:59")

        from_dt = _parse_iso_datetime(from_date)
        to_dt = _parse_iso_datetime(to_date)
        if from_dt is None or to_dt is None:
            raise ValueError("Invalid from_date/to_date format.")

        # 1) Employee names are lightweight and can stay live.
        employees = await self._hub.get_employees(installation_id)
        broker_map: dict[str, str] = {}
        for emp in employees or []:
            eid = emp.get("employeeId")
            name = emp.get("name")
            if eid and name:
                broker_map[str(eid).strip()] = str(name).strip()

        # 2) Sync missing/new data into local cache and build report from cached rows.
        async with async_session_factory() as db:
            try:
                sync_result = await self._sync_sales_cache(
                    db=db,
                    installation_id=installation_id,
                    department_id=department_id,
                    from_dt=from_dt,
                    to_dt=to_dt,
                )
                report_tuple = await self._build_report_tuple_from_cache(
                    db=db,
                    installation_id=installation_id,
                    department_id=department_id,
                    year=y,
                    from_date=from_date,
                    to_date=to_date,
                    from_dt=from_dt,
                    to_dt=to_dt,
                    include_vat=include_vat,
                    broker_map=broker_map,
                    sync_metadata=sync_result,
                )
                await db.commit()
                return report_tuple
            except Exception as cache_err:
                await db.rollback()
                logger.warning("Sales cache sync fallback to live fetch: %s", cache_err)
                return await self._fetch_report_data_live(
                    installation_id=installation_id,
                    department_id=department_id,
                    year=y,
                    from_date=from_date,
                    to_date=to_date,
                    include_vat=include_vat,
                    broker_map=broker_map,
                )

    async def _fetch_report_data_live(
        self,
        *,
        installation_id: str,
        department_id: int,
        year: int,
        from_date: str,
        to_date: str,
        include_vat: bool,
        broker_map: dict[str, str],
    ) -> tuple:
        """Fallback path that fetches live data directly from Vitec."""
        changed_after = from_date[:10] + "T00:00:00"
        estates = await self._hub.get_accounting_estates(
            installation_id,
            department_id=department_id,
            changed_after=changed_after,
        )
        estate_address: dict[str, str] = {}
        estate_metadata: dict[str, dict[str, str]] = {}
        brokers_with_sales: set[str] = set()
        for est in estates or []:
            eid = str(est.get("estateId") or "").strip()
            addr = _build_estate_address(est)
            if eid:
                estate_address[eid] = addr or "(ukjent adresse)"
                estate_metadata[eid] = _build_estate_metadata(est)
            sold_dt = _parse_iso_datetime(est.get("sold"))
            if sold_dt is None or sold_dt.year != year:
                continue
            for b in est.get("brokersIdWithRoles") or []:
                bid = b.get("employeeId")
                if bid:
                    brokers_with_sales.add(str(bid).strip())

        transactions = await self._hub.get_accounting_transactions(
            installation_id,
            from_date=from_date,
            to_date=to_date,
            department_id=department_id,
            ledger_type=1,
        )
        broker_estates: dict[str, dict[str, list[dict]]] = {}
        for txn in transactions or []:
            account = _normalize_account(txn.get("account"))
            if account not in REVENUE_ACCOUNTS:
                continue
            user_id = str(txn.get("userId") or "").strip()
            if not user_id:
                continue
            if brokers_with_sales and user_id not in brokers_with_sales:
                continue
            estate_id = str(txn.get("estateId") or "").strip() or "(ukjent)"
            broker_estates.setdefault(user_id, {}).setdefault(estate_id, []).append(txn)

        if not brokers_with_sales and broker_estates:
            brokers_with_sales = set(broker_estates.keys())

        report_data = self._build_report_data_dict(
            year=year,
            department_id=department_id,
            from_date=from_date,
            to_date=to_date,
            include_vat=include_vat,
            broker_map=broker_map,
            brokers_with_sales=brokers_with_sales,
            broker_estates=broker_estates,
            estate_address=estate_address,
            estate_metadata=estate_metadata,
        )
        return report_data, broker_map, brokers_with_sales, broker_estates, estate_address, estate_metadata, include_vat

    async def _sync_sales_cache(
        self,
        *,
        db,
        installation_id: str,
        department_id: int,
        from_dt: datetime,
        to_dt: datetime,
    ) -> dict[str, Any]:
        # Advisory lock prevents concurrent syncs for the same department.
        lock_key = hash(f"{installation_id}:{department_id}") & 0x7FFFFFFF
        await db.execute(text(f"SELECT pg_advisory_xact_lock({lock_key})"))

        state_key = f"{installation_id}:{department_id}"
        state = await db.get(ReportSalesCacheState, state_key)
        if state is None:
            state = ReportSalesCacheState(
                state_key=state_key,
                installation_id=installation_id,
                department_id=department_id,
                month_sync_json={},
            )
            db.add(state)
            await db.flush()

        estates_upserted = 0
        transactions_upserted = 0
        skipped: dict[str, int] = {"future_date": 0, "orphan": 0}
        warnings_count = 0
        now_utc = datetime.now(timezone.utc)
        tomorrow = now_utc + timedelta(days=1)

        # Estates: incremental by changedAfter cursor (+24h overlap).
        changed_after_dt = state.last_estates_sync_at - timedelta(days=1) if state.last_estates_sync_at else from_dt
        changed_after = changed_after_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
        estates = await self._hub.get_accounting_estates(
            installation_id,
            department_id=department_id,
            changed_after=changed_after,
        )
        for est in estates or []:
            estate_id = str(est.get("estateId") or "").strip()
            if not estate_id:
                continue
            estate_key = f"{installation_id}:{estate_id}"
            row = await db.get(ReportSalesEstateCache, estate_key)
            metadata = _build_estate_metadata(est)
            sold_dt = _parse_iso_datetime(est.get("sold"))
            payload_brokers = est.get("brokersIdWithRoles") or []
            oppdrag = metadata.get("assignment_number") or None
            if row is None:
                row = ReportSalesEstateCache(
                    estate_key=estate_key,
                    installation_id=installation_id,
                    department_id=department_id,
                    estate_id=estate_id,
                    data_source=DATA_SOURCE_VITEC_NEXT,
                    sold_at=sold_dt,
                    address=_build_estate_address(est) or "(ukjent adresse)",
                    property_type=metadata["property_type"],
                    assignment_type=metadata["assignment_type"],
                    assignment_number=oppdrag,
                    brokers_json=payload_brokers,
                )
                db.add(row)
                estates_upserted += 1
            else:
                row.department_id = department_id
                row.data_source = DATA_SOURCE_VITEC_NEXT
                row.sold_at = sold_dt
                row.address = _build_estate_address(est) or "(ukjent adresse)"
                row.property_type = metadata["property_type"]
                row.assignment_type = metadata["assignment_type"]
                row.assignment_number = oppdrag
                row.brokers_json = payload_brokers
                estates_upserted += 1

        # Transactions: sync unsynced months, always refresh current month.
        current_month_key = _month_key(now_utc.year, now_utc.month)
        month_sync = dict(state.month_sync_json or {})
        months = _iter_months(from_dt, to_dt)
        for year, month in months:
            mk = _month_key(year, month)
            should_fetch = mk not in month_sync or mk == current_month_key
            if not should_fetch:
                continue

            month_start, month_end = _month_bounds(year, month)
            txns = await self._hub.get_accounting_transactions(
                installation_id,
                from_date=month_start.strftime("%Y-%m-%dT%H:%M:%S"),
                to_date=month_end.strftime("%Y-%m-%dT%H:%M:%S"),
                department_id=department_id,
                ledger_type=1,
            )
            for txn in txns or []:
                posting_dt = _parse_iso_datetime(txn.get("postingDate"))

                # Validation: reject future-dated transactions
                if posting_dt and posting_dt > tomorrow:
                    skipped["future_date"] += 1
                    continue

                # Validation: skip orphan records with no estate AND no user
                user_id = str(txn.get("userId") or "").strip()
                estate_id_raw = str(txn.get("estateId") or "").strip()
                if not user_id and not estate_id_raw:
                    skipped["orphan"] += 1
                    continue

                # Anomaly warning: unusually large amounts
                amt = abs(float(txn.get("amount") or 0))
                if amt > 5_000_000:
                    warnings_count += 1
                    logger.warning(
                        "Large transaction amount %.2f for dept %s estate %s",
                        amt,
                        department_id,
                        estate_id_raw,
                    )

                tx_key = self._transaction_cache_key(installation_id=installation_id, txn=txn)
                row = await db.get(ReportSalesTransactionCache, tx_key)
                if row is None:
                    row = ReportSalesTransactionCache(
                        transaction_key=tx_key,
                        installation_id=installation_id,
                        department_id=department_id,
                        data_source=DATA_SOURCE_VITEC_NEXT,
                        posting_date=posting_dt,
                        account=str(txn.get("account") or "").strip(),
                        user_id=user_id,
                        estate_id=estate_id_raw,
                        amount=float(txn.get("amount") or 0),
                        vat_amount=float(txn.get("vatAmount") or 0),
                        description=str(txn.get("description") or ""),
                    )
                    db.add(row)
                    transactions_upserted += 1
                else:
                    row.department_id = department_id
                    row.data_source = DATA_SOURCE_VITEC_NEXT
                    row.posting_date = posting_dt
                    row.account = str(txn.get("account") or "").strip()
                    row.user_id = user_id
                    row.estate_id = estate_id_raw
                    row.amount = float(txn.get("amount") or 0)
                    row.vat_amount = float(txn.get("vatAmount") or 0)
                    row.description = str(txn.get("description") or "")
                    transactions_upserted += 1
            month_sync[mk] = now_utc.isoformat()

        state.month_sync_json = month_sync
        state.last_estates_sync_at = now_utc
        event = ReportSalesSyncEvent(
            installation_id=installation_id,
            department_id=department_id,
            event_type="cache_sync",
            from_date=from_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            to_date=to_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            estates_upserted=estates_upserted,
            transactions_upserted=transactions_upserted,
            payload_json={
                "months_scanned": len(months),
                "current_month_refresh": current_month_key,
                "rows_ingested": estates_upserted + transactions_upserted,
                "rows_skipped": skipped,
                "validation_warnings_count": warnings_count,
            },
        )
        db.add(event)
        return {
            "estates_upserted": estates_upserted,
            "transactions_upserted": transactions_upserted,
            "last_synced_at": now_utc.isoformat(),
            "validation_warnings_count": warnings_count,
        }

    async def _build_report_tuple_from_cache(
        self,
        *,
        db,
        installation_id: str,
        department_id: int,
        year: int,
        from_date: str,
        to_date: str,
        from_dt: datetime,
        to_dt: datetime,
        include_vat: bool,
        broker_map: dict[str, str],
        sync_metadata: dict[str, Any] | None = None,
    ) -> tuple:
        estates_stmt = select(ReportSalesEstateCache).where(
            and_(
                ReportSalesEstateCache.installation_id == installation_id,
                ReportSalesEstateCache.department_id == department_id,
            )
        )
        estates = list((await db.execute(estates_stmt)).scalars().all())

        estate_address: dict[str, str] = {}
        estate_metadata: dict[str, dict[str, str]] = {}
        brokers_with_sales: set[str] = set()
        for est in estates:
            eid = str(est.estate_id or "").strip()
            if not eid:
                continue
            estate_address[eid] = est.address or "(ukjent adresse)"
            estate_metadata[eid] = {
                "property_type": est.property_type or "—",
                "assignment_type": est.assignment_type or "—",
                "assignment_number": est.assignment_number or "",
            }
            sold_dt = est.sold_at
            if sold_dt is None or sold_dt.year != year:
                continue
            for b in est.brokers_json or []:
                bid = str((b or {}).get("employeeId") or "").strip()
                if bid:
                    brokers_with_sales.add(bid)

        tx_stmt = select(ReportSalesTransactionCache).where(
            and_(
                ReportSalesTransactionCache.installation_id == installation_id,
                ReportSalesTransactionCache.department_id == department_id,
                ReportSalesTransactionCache.posting_date.is_not(None),
                ReportSalesTransactionCache.posting_date >= from_dt,
                ReportSalesTransactionCache.posting_date <= to_dt,
            )
        )
        tx_rows = list((await db.execute(tx_stmt)).scalars().all())

        broker_estates: dict[str, dict[str, list[dict[str, Any]]]] = {}
        for txn in tx_rows:
            account = _normalize_account(txn.account)
            if account not in REVENUE_ACCOUNTS:
                continue
            user_id = str(txn.user_id or "").strip()
            if not user_id:
                continue
            if brokers_with_sales and user_id not in brokers_with_sales:
                continue
            estate_id = str(txn.estate_id or "").strip() or "(ukjent)"
            txn_payload = {
                "account": txn.account,
                "userId": user_id,
                "estateId": estate_id,
                "amount": txn.amount,
                "vatAmount": txn.vat_amount,
                "description": txn.description,
                "postingDate": txn.posting_date.isoformat() if txn.posting_date else None,
            }
            broker_estates.setdefault(user_id, {}).setdefault(estate_id, []).append(txn_payload)

        if not brokers_with_sales and broker_estates:
            brokers_with_sales = set(broker_estates.keys())

        # Count rows by data source for scope metadata
        source_counts: dict[str, int] = {}
        src_stmt = (
            select(
                ReportSalesTransactionCache.data_source,
                sa_func.count().label("cnt"),
            )
            .where(
                and_(
                    ReportSalesTransactionCache.installation_id == installation_id,
                    ReportSalesTransactionCache.department_id == department_id,
                    ReportSalesTransactionCache.posting_date.is_not(None),
                    ReportSalesTransactionCache.posting_date >= from_dt,
                    ReportSalesTransactionCache.posting_date <= to_dt,
                )
            )
            .group_by(ReportSalesTransactionCache.data_source)
        )
        for row in (await db.execute(src_stmt)).all():
            source_counts[row[0]] = row[1]

        report_data = self._build_report_data_dict(
            year=year,
            department_id=department_id,
            from_date=from_date,
            to_date=to_date,
            include_vat=include_vat,
            broker_map=broker_map,
            brokers_with_sales=brokers_with_sales,
            broker_estates=broker_estates,
            estate_address=estate_address,
            estate_metadata=estate_metadata,
            sync_metadata=sync_metadata,
            source_counts=source_counts,
        )

        return report_data, broker_map, brokers_with_sales, broker_estates, estate_address, estate_metadata, include_vat

    @staticmethod
    async def _resolve_department_names(department_ids: list[int]) -> dict[int, str]:
        """Map Vitec department IDs to office names from the offices table."""
        if not department_ids:
            return {}
        try:
            async with async_session_factory() as db:
                stmt = select(Office.vitec_department_id, Office.name).where(
                    Office.vitec_department_id.in_(department_ids),
                )
                rows = (await db.execute(stmt)).all()
                return {int(r[0]): r[1] for r in rows if r[0] is not None}
        except Exception:
            logger.debug("Could not resolve department names", exc_info=True)
            return {}

    @staticmethod
    def _transaction_cache_key(*, installation_id: str, txn: dict) -> str:
        raw = "|".join(
            [
                installation_id,
                str(txn.get("postingDate") or ""),
                str(txn.get("account") or ""),
                str(txn.get("userId") or ""),
                str(txn.get("estateId") or ""),
                str(txn.get("amount") or ""),
                str(txn.get("vatAmount") or ""),
                str(txn.get("description") or ""),
            ]
        )
        return hashlib.sha1(raw.encode("utf-8")).hexdigest()

    def _build_report_data_dict(
        self,
        *,
        year: int,
        department_id: int,
        from_date: str,
        to_date: str,
        include_vat: bool,
        broker_map: dict[str, str],
        brokers_with_sales: set[str],
        broker_estates: dict[str, dict[str, list[dict]]],
        estate_address: dict[str, str],
        estate_metadata: dict[str, dict[str, str]],
        sync_metadata: dict[str, Any] | None = None,
        source_counts: dict[str, int] | None = None,
    ) -> dict:
        report_data = {
            "year": year,
            "department_id": department_id,
            "from_date": from_date,
            "to_date": to_date,
            "from_date_display": _format_date_iso(from_date) or from_date[:10],
            "to_date_display": _format_date_iso(to_date) or to_date[:10],
            "include_vat": include_vat,
            "brokers": [],
        }

        def _display_address(eid: str, addr: str, meta: dict[str, str]) -> str:
            """Show address or 'Adresse ukjent' + oppdragsnummer, never raw estate_id (UUID)."""
            oppdrag = meta.get("assignment_number") or ""
            if addr and not _looks_like_uuid(addr):
                return f"{addr} ({oppdrag})" if oppdrag else addr
            suffix = f" ({oppdrag})" if oppdrag else ""
            return f"Adresse ukjent{suffix}" if suffix else "Adresse ukjent"

        for broker_id in sorted(brokers_with_sales):
            name = broker_map.get(broker_id, broker_id)
            estates_data = broker_estates.get(broker_id, {})
            sale_count = len(estates_data)
            total = 0.0
            properties: list[dict] = []
            for estate_id, txns in sorted(estates_data.items(), key=lambda x: estate_address.get(x[0], x[0])):
                raw_addr = estate_address.get(estate_id, estate_id)
                meta = estate_metadata.get(
                    estate_id,
                    {"property_type": "—", "assignment_type": "—", "assignment_number": ""},
                )
                addr = _display_address(estate_id, raw_addr, meta)
                prop_total = 0.0
                txns_data: list[dict] = []
                for txn in txns:
                    amt = _revenue_amount(
                        float(txn.get("amount") or 0),
                        float(txn.get("vatAmount") or 0),
                        include_vat,
                    )
                    prop_total += amt
                    total += amt
                    txns_data.append(
                        {
                            "posting_date": _format_date_iso(txn.get("postingDate")),
                            "account": txn.get("account"),
                            "description": (txn.get("description") or "")[:80],
                            "amount": round(amt, 2),
                        }
                    )
                properties.append(
                    {
                        "address": addr,
                        "estate_id": estate_id,
                        "assignment_number": meta.get("assignment_number") or "",
                        "property_type": meta["property_type"],
                        "assignment_type": meta["assignment_type"],
                        "total": round(prop_total, 2),
                        "transactions": txns_data,
                    }
                )
            report_data["brokers"].append(
                {
                    "broker_id": broker_id,
                    "name": name,
                    "sale_count": sale_count,
                    "total": round(total, 2),
                    "properties": properties,
                }
            )

        report_data["brokers"] = [b for b in report_data["brokers"] if b["sale_count"] > 0]
        report_data["total_sales"] = sum(b["sale_count"] for b in report_data["brokers"])
        report_data["total_revenue"] = round(sum(b["total"] for b in report_data["brokers"]), 2)

        data_sources = []
        sc = source_counts or {}
        for src_name, src_cfg in SYNC_SOURCES.items():
            data_sources.append(
                {
                    "name": src_name,
                    "label": src_cfg["label"],
                    "coverage": f"{src_cfg['coverage_start']} - present",
                    "row_count": sc.get(src_name, 0),
                }
            )

        report_data["scope"] = {
            "accounts_included": sorted(REVENUE_ACCOUNTS),
            "account_categories": {
                "vederlag": sorted(VEDERLAG_ACCOUNTS),
                "andre_inntekter": sorted(ANDRE_INNTEKTER_ACCOUNTS),
            },
            "estate_statuses": "40-48 (solgt/overtatt)",
            "vat_handling": "included" if include_vat else "excluded",
            "date_range": {"from": from_date, "to": to_date},
            "department_filter": department_id,
            "last_synced_at": (sync_metadata or {}).get("last_synced_at"),
            "data_sources": data_sources,
            "brokers_filter": "only brokers with sales in period",
            "data_freshness_note": "Current month re-synced on every load; past months cached",
            "validation_warnings_count": (sync_metadata or {}).get("validation_warnings_count", 0),
        }

        return report_data

    def _build_excel(
        self,
        *,
        year: int,
        department_id: int,
        from_date: str,
        to_date: str,
        broker_map: dict[str, str],
        brokers_with_sales: set[str],
        broker_estates: dict[str, dict[str, list[dict]]],
        estate_address: dict[str, str],
        estate_metadata: dict[str, dict[str, str]],
        include_vat: bool,
    ) -> bytes:
        """Generate Excel workbook with expandable broker/property/transaction detail."""
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formidlingsrapport"

        from_display = _format_date_iso(from_date) or from_date[:10]
        to_display = _format_date_iso(to_date) or to_date[:10]
        sum_label = "Sum vederlag + andre inntekter (kr)" + (" inkl. mva." if include_vat else " eksl. mva.")

        # Header
        ws.append(["Formidlingsrapport - Vederlag og andre inntekter"])
        ws.merge_cells("A1:E1")
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Avdeling: {department_id}  |  År: {year}  |  Periode: {from_display} - {to_display}"])
        ws.merge_cells("A2:E2")
        ws.append([])

        # Column headers: Megler, Antall salg, Eiendomstype, Oppdragstype, Sum
        headers = ["Megler", "Antall salg", "Eiendomstype", "Oppdragstype", sum_label]
        ws.append(headers)
        header_row = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=header_row, column=col).font = Font(bold=True)

        # Data rows with expandable detail
        # Only include brokers with at least one sale
        brokers_sorted = [b for b in sorted(brokers_with_sales) if len(broker_estates.get(b, {})) > 0]
        row = header_row + 1

        for broker_id in brokers_sorted:
            name = broker_map.get(broker_id, broker_id)
            estates = broker_estates.get(broker_id, {})
            sale_count = len(estates)

            total = 0.0
            for estate_txns in estates.values():
                for txn in estate_txns:
                    total += _revenue_amount(
                        float(txn.get("amount") or 0),
                        float(txn.get("vatAmount") or 0),
                        include_vat,
                    )

            # Broker summary row (level 0) - no property/assignment type at broker level
            ws.append([name, sale_count, "", "", round(total, 0)])
            broker_row = row
            row += 1

            # Property and transaction detail (grouped)
            for estate_id, txns in sorted(estates.items(), key=lambda x: estate_address.get(x[0], x[0])):
                raw_addr = estate_address.get(estate_id, estate_id)
                meta = estate_metadata.get(
                    estate_id,
                    {"property_type": "—", "assignment_type": "—", "assignment_number": ""},
                )
                oppdrag = meta.get("assignment_number") or ""
                if raw_addr and not _looks_like_uuid(raw_addr):
                    addr = f"{raw_addr} ({oppdrag})" if oppdrag else raw_addr
                else:
                    suffix = f" ({oppdrag})" if oppdrag else ""
                    addr = f"Adresse ukjent{suffix}" if suffix else "Adresse ukjent"
                # Property row (level 1)
                ws.append([f"  {addr}", "", meta["property_type"], meta["assignment_type"], ""])
                ws.cell(row=row, column=1).font = Font(italic=True)
                prop_row = row
                row += 1

                # Transaction rows (level 2)
                for txn in txns:
                    amt = _revenue_amount(
                        float(txn.get("amount") or 0),
                        float(txn.get("vatAmount") or 0),
                        include_vat,
                    )
                    post_date = _format_date_iso(txn.get("postingDate"))
                    acc = txn.get("account") or ""
                    desc = (txn.get("description") or "")[:40]
                    ws.append([f"    {post_date} | Konto {acc} | {desc}", "", "", "", round(amt, 0)])
                    row += 1

                # Group property + its transactions
                ws.row_dimensions.group(prop_row, row - 1, outline_level=2, hidden=False)

            # Group broker + all properties
            if row > broker_row + 1:
                ws.row_dimensions.group(broker_row + 1, row - 1, outline_level=1, hidden=False)

        # Total row
        total_revenue = sum(
            sum(
                _revenue_amount(float(t.get("amount") or 0), float(t.get("vatAmount") or 0), include_vat)
                for est_txns in broker_estates.get(b, {}).values()
                for t in est_txns
            )
            for b in brokers_sorted
        )
        total_sales = sum(len(broker_estates.get(b, {})) for b in brokers_sorted)
        ws.append(["Sum", total_sales, "", "", round(total_revenue, 0)])
        total_row = row
        for col in range(1, 6):
            ws.cell(row=total_row, column=col).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 38

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
